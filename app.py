from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, traceback
from datetime import datetime

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

def fill(h):    return PatternFill("solid", fgColor=h)
def border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)
def ca():       return Alignment(horizontal="center", vertical="center", wrap_text=True)
def la():       return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def ra():       return Alignment(horizontal="right",  vertical="center")
def hf(sz=10, bold=True, color="FFFFFF"):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def bf(sz=10, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)
def inp():      return Font(name="Arial", size=10, color="0000FF")
def xf():       return Font(name="Arial", size=10, color="006400")

FMT_INR = '₹#,##0;(₹#,##0);"-"'
FMT_AED = '"AED "#,##0.00;("AED "#,##0.00);"-"'
FMT_PCT = '0.0%'
FMT_0   = '0.00x'

def sc(ws, row, col, val, fmt="General", fnt=None, bg="FFFFFF", aln=None, brd=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = fnt or bf(); cell.fill = fill(bg)
    cell.number_format = fmt; cell.alignment = aln or la()
    if brd: cell.border = border()
    return cell

def merge_sc(ws, row, c1, c2, val, fmt="General", fnt=None, bg="FFFFFF", aln=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    return sc(ws, row, c1, val, fmt, fnt, bg, aln)

def banner(ws, row, text, bg, end_col=9, sz=13, h=44):
    merge_sc(ws, row, 1, end_col, text, fnt=hf(sz=sz), bg=bg, aln=ca())
    ws.row_dimensions[row].height = h

def section(ws, row, text, bg, end_col=9):
    merge_sc(ws, row, 1, end_col, text, fnt=hf(sz=10), bg=bg, aln=ca())
    ws.row_dimensions[row].height = 22

def hdr(ws, row, cols, bg="1A2B4A", start=1):
    for i, txt in enumerate(cols, start):
        cell = ws.cell(row=row, column=i, value=txt)
        cell.font = hf(sz=9); cell.fill = fill(bg)
        cell.alignment = ca(); cell.border = border()
    ws.row_dimensions[row].height = 26

def cw(ws, col, w): ws.column_dimensions[col].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h


def build_tracker(data):
    def sf(v, default=0.0):
        """Safe float — returns default if v is empty string or None."""
        try:
            return float(v) if v not in (None, '', ' ') else default
        except (ValueError, TypeError):
            return default

    wb = Workbook()
    wb.remove(wb.active)

    name        = data.get("name", "User")
    spouse      = data.get("spouseName", "")
    currency    = data.get("currency", "INR")
    country     = data.get("country", "India")
    income      = sf(data.get("income", 0))
    spouse_inc  = sf(data.get("spouseIncome", 0))
    investments = data.get("investments", [])
    liabilities = data.get("liabilities", [])
    policies    = data.get("policies", [])
    goals       = data.get("goals", [])
    risk        = data.get("riskAppetite", "Medium")
    is_nri      = country.lower() not in ("india",)

    SET  = "'⚙ Settings'"
    INV  = "'📈 Investments'"
    LIAB = "'📋 Liabilities'"
    MC   = "'📅 Monthly Commitments'"
    POL  = "'🛡 Policies'"
    SAL  = "'💡 Salary Allocation'"

    # ── SHEET 1: Settings ─────────────────────────────────────────
    ws = wb.create_sheet("⚙ Settings")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1A2B4A"
    banner(ws, 1, "⚙  GLOBAL SETTINGS", "1A2B4A", 5, sz=13)

    rows = [
        (3, "Last Updated Date:",  "=TODAY()",  "DD-MMM-YYYY", False),
        (4, "AED → INR Live Rate:",
         '=IFERROR(__xludf.dummyfunction("ROUND(IFERROR(INDEX(googlefinance(\\"CURRENCY:AEDINR\\",\\"close\\",TODAY()-1),2,2),25.5),4)"),25.5)',
         "0.0000", False),
        (5, f"{name} Income ({currency}/month):", income,      FMT_INR, True),
        (6, f"{spouse or 'Spouse'} Income ({currency}/month):", spouse_inc, FMT_INR, True),
        (7, "Combined Income (INR/month):",
         f"=(B5+B6)*B4" if currency == "AED" else "=B5+B6",
         FMT_INR, False),
    ]
    for row, label, val, fmt, is_inp in rows:
        sc(ws, row, 1, label, fnt=bf(bold=True), aln=ra())
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sc(ws, row, 3, val, fmt=fmt,
           fnt=inp() if is_inp else xf(),
           bg="FEF9E7" if is_inp else "D6E4F0", aln=ra())
        sc(ws, row, 4, "← update when salary changes" if is_inp else "← auto", fnt=bf(sz=8, color="888888"), aln=la())
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        rh(ws, row, 20)

    merge_sc(ws, 9, 1, 5,
        f"Tracker for: {name}" + (f" & {spouse}" if spouse else "") +
        f"  |  {country}  |  Risk: {risk}  |  Currency: {currency}",
        fnt=bf(sz=9, color="555555"), bg="F2F2F2", aln=la())
    for col, w in [("A",24),("B",18),("C",16),("D",30),("E",10)]:
        cw(ws, col, w)

    # ── SHEET 2: Investments ──────────────────────────────────────
    ws = wb.create_sheet("📈 Investments")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "27AE60"
    ws.freeze_panes = "A4"
    banner(ws, 1, "📈  INVESTMENTS — Update Current Value monthly", "1A2B4A", 13)
    merge_sc(ws, 2, 1, 13,
        "🔵 Blue = you type  |  🟢 Green = auto-calculated  |  Add a row for each investment",
        fnt=bf(sz=9, color="555555"), aln=la())

    inv_hdrs = ["Investment Name","Category","Invested (INR)","Invested (AED)",
                "Invested (Equiv)","Cur Value (INR)","Cur Value (AED)",
                "Cur Value (Equiv)","Status","Belongs To","Recurring?","Life Cover?","Notes"]
    hdr(ws, 3, inv_hdrs)

    cat_map = {
        "Mutual Fund":"Mutual Fund","MF":"Mutual Fund","Index Fund":"Mutual Fund",
        "NPS":"NPS","PPF":"PPF","FD":"Post Office","RD":"Post Office",
        "Gold":"Gold","Stocks":"Stocks India","ETF":"Stocks UAE",
        "IBKR":"Stocks UAE","Real Estate":"Real Estate UAE",
        "Emergency":"Savings","SSY":"SSY","LIC":"LIC Policy","Insurance":"LIC Policy",
        "Other":"Other",
    }

    for i, inv in enumerate(investments):
        r_num = 4 + i
        bg = "F2F2F2" if r_num % 2 == 1 else "FFFFFF"
        cat_raw = inv.get("type","Other")
        cat = cat_map.get(cat_raw, cat_raw)
        vals = [
            inv.get("name",""),   cat,
            sf(inv.get("investedINR",0)), sf(inv.get("investedAED",0)),
            f"=C{r_num}+D{r_num}*{SET}!$B$4",
            sf(inv.get("currentINR",0)), sf(inv.get("currentAED",0)),
            f"=F{r_num}+G{r_num}*{SET}!$B$4",
            "Active", inv.get("owner", name),
            "Yes" if inv.get("recurring") else "No", "N",
            inv.get("notes",""),
        ]
        fmts = ["","","",FMT_INR,"",FMT_INR,"",FMT_INR,"","","","",""]
        for j, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r_num, j, value=v)
            cell.fill = fill("D6E4F0" if j in (5,8) else bg)
            cell.border = border(); cell.number_format = fmt or "General"
            if j in (5,8):   cell.font = xf(); cell.alignment = ra()
            elif j in (3,4,6,7): cell.font = inp(); cell.alignment = ra()
            elif j == 1:     cell.font = bf(bold=True); cell.alignment = la()
            else:            cell.font = bf(sz=9); cell.alignment = ca() if j < 13 else la()

    for col, w in [("A",30),("B",16),("C",14),("D",13),("E",17),("F",14),
                   ("G",13),("H",17),("I",12),("J",12),("K",10),("L",10),("M",35)]:
        cw(ws, col, w)

    # ── SHEET 3: Liabilities ──────────────────────────────────────
    ws = wb.create_sheet("📋 Liabilities")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "E74C3C"
    ws.freeze_panes = "A4"
    banner(ws, 1, "📋  LIABILITIES — Update Outstanding monthly", "E74C3C", 11)

    liab_hdrs = ["Loan","Bank","Type","Outstanding (INR)","Outstanding (AED)",
                 "Outstanding (Equiv)","Interest %","EMI (INR)","EMI (AED)",
                 "EMI (Equiv)","Notes"]
    hdr(ws, 3, liab_hdrs)

    for i, liab in enumerate(liabilities):
        r_num = 4 + i
        bg = "F2F2F2" if r_num % 2 == 1 else "FFFFFF"
        vals = [
            liab.get("name",""), liab.get("bank",""), liab.get("type","Personal Loan"),
            sf(liab.get("outstandingINR",0)), sf(liab.get("outstandingAED",0)),
            f"=D{r_num}+E{r_num}*{SET}!$B$4",
            sf(liab.get("rate",0))/100,
            sf(liab.get("emiINR",0)), sf(liab.get("emiAED",0)),
            f"=H{r_num}+I{r_num}*{SET}!$B$4",
            liab.get("notes",""),
        ]
        fmts = ["","","",FMT_INR,FMT_AED,FMT_INR,FMT_PCT,FMT_INR,FMT_AED,FMT_INR,""]
        for j, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r_num, j, value=v)
            cell.fill = fill("D6E4F0" if j in (6,10) else bg)
            cell.border = border(); cell.number_format = fmt or "General"
            if j in (6,10):   cell.font = xf(); cell.alignment = ra()
            elif j in (4,5,7,8,9): cell.font = inp(); cell.alignment = ra()
            elif j == 1:      cell.font = bf(bold=True); cell.alignment = la()
            else:             cell.font = bf(sz=9); cell.alignment = la()

    liab_total_r = 4 + len(liabilities) + 1
    ws.cell(liab_total_r,1).value = "TOTAL"
    ws.cell(liab_total_r,1).font = hf(sz=10)
    ws.cell(liab_total_r,1).fill = fill("E74C3C")
    ws.cell(liab_total_r,1).alignment = ca(); ws.cell(liab_total_r,1).border = border()
    for col, fmt in [(4,FMT_INR),(6,FMT_INR),(8,FMT_INR),(10,FMT_INR)]:
        col_l = get_column_letter(col)
        cell = ws.cell(liab_total_r, col, value=f"=SUM({col_l}4:{col_l}{liab_total_r-1})")
        cell.font = hf(sz=10); cell.fill = fill("E74C3C")
        cell.number_format = fmt; cell.alignment = ra(); cell.border = border()

    for col, w in [("A",26),("B",14),("C",14),("D",16),("E",14),("F",18),
                   ("G",12),("H",16),("I",14),("J",18),("K",35)]:
        cw(ws, col, w)

    # ── SHEET 4: Assets (auto) ────────────────────────────────────
    ws = wb.create_sheet("💰 Assets")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F39C12"
    banner(ws, 1, "💰  ASSETS SUMMARY — Auto-calculated from 📈 Investments", "1A2B4A", 8)
    merge_sc(ws, 2, 1, 8,
        "ℹ️  Do not edit here — update 📈 Investments instead. This sheet is 100% formula-driven.",
        fnt=bf(sz=9, color="555555"), aln=la())

    hdr(ws, 4, ["Asset Class","Invested (INR)","Invested (AED)","Invested (Equiv)",
                "Cur Value (INR)","Cur Value (AED)","Cur Value (Equiv)","% of Total"])

    asset_cats = [
        ("Mutual Funds","Mutual Fund"),("NPS","NPS"),("PPF","PPF"),
        ("KSFE Chitty","KSFE Chitty"),("Chitty/Society","Chitty/Society"),
        ("LIC Policies","LIC Policy"),("Term Insurance","Term Insurance"),
        ("Post Office","Post Office"),("Gold","Gold"),
        ("Stocks (India)","Stocks India"),("Stocks (UAE)","Stocks UAE"),
        ("Real Estate (UAE)","Real Estate UAE"),
        ("Savings / Emergency Fund","Savings"),
        ("SSY (Girl Child)","SSY"),("Govt Pension","Govt Pension"),("Other","Other"),
    ]

    data_start = 5
    for i, (label, cat) in enumerate(asset_cats):
        r_num = data_start + i
        bg = "F2F2F2" if i % 2 == 1 else "FFFFFF"
        gt_assets = data_start + len(asset_cats) + 1
        ws.cell(r_num,1).value = label
        ws.cell(r_num,1).font = bf(bold=True); ws.cell(r_num,1).fill = fill(bg)
        ws.cell(r_num,1).border = border(); ws.cell(r_num,1).alignment = la()
        for col, src, fmt in [
            (2,"$C:$C",FMT_INR),(3,"$D:$D",FMT_AED),(4,"$E:$E",FMT_INR),
            (5,"$F:$F",FMT_INR),(6,"$G:$G",FMT_AED),(7,"$H:$H",FMT_INR),
        ]:
            cell = ws.cell(r_num, col, value=f"=SUMIF({INV}!$B:$B,\"{cat}\",{INV}!{src})")
            cell.font = xf(); cell.fill = fill("D6E4F0" if col in (4,7) else bg)
            cell.number_format = fmt; cell.alignment = ra(); cell.border = border()
        cell = ws.cell(r_num, 8, value=f"=IFERROR(G{r_num}/G{gt_assets},0)")
        cell.font = bf(); cell.fill = fill(bg); cell.number_format = FMT_PCT
        cell.alignment = ra(); cell.border = border()

    gt_assets = data_start + len(asset_cats) + 1
    ws.cell(gt_assets,1).value = "GRAND TOTAL"
    ws.cell(gt_assets,1).font = hf(sz=10); ws.cell(gt_assets,1).fill = fill("1A2B4A")
    ws.cell(gt_assets,1).alignment = ca(); ws.cell(gt_assets,1).border = border()
    for col, fmt in [(2,FMT_INR),(3,FMT_AED),(4,FMT_INR),(5,FMT_INR),(6,FMT_AED),(7,FMT_INR),(8,"0%")]:
        col_l = get_column_letter(col)
        cell = ws.cell(gt_assets, col)
        cell.value = f"=SUM({col_l}{data_start}:{col_l}{gt_assets-1})" if col != 8 else "=1"
        cell.font = hf(sz=10); cell.fill = fill("1A2B4A")
        cell.number_format = fmt; cell.alignment = ra(); cell.border = border()

    for col, w in [("A",22),("B",16),("C",14),("D",18),("E",16),("F",14),("G",18),("H",10)]:
        cw(ws, col, w)

    # ── SHEET 5: Policies ─────────────────────────────────────────
    ws = wb.create_sheet("🛡 Policies")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "8E44AD"
    banner(ws, 1, "🛡  INSURANCE POLICIES", "4A235A", 10)
    hdr(ws, 2, ["Policy Type","Policy #","Sum Assured","Death Cover","Status",
                "Holder","Life Cover?","Premium (INR/month)","Agent","Notes"])

    for i, pol in enumerate(policies):
        r_num = 3 + i
        bg = "F2F2F2" if r_num % 2 == 1 else "FFFFFF"
        vals = [pol.get("type","Term"), pol.get("policyNum",""),
                sf(pol.get("sumAssured",0)), sf(pol.get("deathCover",0)),
                pol.get("status","Active"), pol.get("holder", name),
                "Y" if pol.get("lifeCover") else "N",
                sf(pol.get("premium",0)), pol.get("agent",""), pol.get("notes","")]
        fmts = ["","",FMT_INR,FMT_INR,"","","",FMT_INR,"",""]
        for j, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r_num, j, value=v)
            cell.fill = fill(bg); cell.border = border()
            cell.number_format = fmt or "General"
            if j in (3,4,8): cell.font = inp(); cell.alignment = ra()
            elif j == 1:     cell.font = bf(bold=True); cell.alignment = la()
            else:            cell.font = bf(sz=9); cell.alignment = ca()

    for col, w in [("A",16),("B",20),("C",18),("D",18),("E",14),("F",12),
                   ("G",10),("H",18),("I",14),("J",35)]:
        cw(ws, col, w)

    # ── SHEET 6: Monthly Commitments ─────────────────────────────
    ws = wb.create_sheet("📅 Monthly Commitments")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2D5986"
    ws.freeze_panes = "A4"
    banner(ws, 1, "📅  MONTHLY COMMITMENTS — Tick off payments each month", "2D5986", 10)
    hdr(ws, 3, ["Commitment","Paid By","Type","Amount (INR)","Amount (AED)",
                "INR Equiv","Category","Frequency","Ends By","Months Left","Notes"])

    mc_rows = []
    for inv in investments:
        amt = sf(inv.get("monthlyAmount",0))
        if inv.get("recurring") and amt > 0:
            mc_rows.append({"name":inv.get("name",""),"by":inv.get("owner",name),
                "type":"Investment","inr":amt,"aed":0,"cat":"Investment",
                "freq":"Monthly","ends":inv.get("ends","Ongoing"),"notes":""})
    for liab in liabilities:
        emi = sf(liab.get("emiINR",0))
        if emi > 0:
            mc_rows.append({"name":liab.get("name",""),"by":name,
                "type":"Liability","inr":emi,"aed":0,"cat":"Liability",
                "freq":"Monthly","ends":liab.get("ends",""),"notes":f"{liab.get('rate',0)}% interest"})
    for pol in policies:
        prem = sf(pol.get("premium",0))
        if prem > 0:
            mc_rows.append({"name":f"{pol.get('type','')} Premium","by":pol.get("holder",name),
                "type":"Insurance","inr":prem,"aed":0,"cat":"Insurance",
                "freq":"Monthly","ends":"Policy term","notes":""})

    # Add user-entered monthly commitments from form
    for commit in data.get('commitments', []):
        amt_inr = sf(commit.get('inr', 0))
        amt_aed = sf(commit.get('aed', 0))
        if amt_inr > 0 or amt_aed > 0:
            mc_rows.append({
                'name':  commit.get('name',''),
                'by':    commit.get('by', name),
                'type':  commit.get('type','Investment'),
                'inr':   amt_inr,
                'aed':   amt_aed,
                'cat':   commit.get('cat','Investment'),
                'freq':  commit.get('freq','Monthly'),
                'ends':  'Ongoing',
                'notes': commit.get('notes',''),
            })

    for i, row_d in enumerate(mc_rows):
        r_num = 4 + i
        bg = "F2F2F2" if r_num % 2 == 1 else "FFFFFF"
        vals = [row_d["name"],row_d["by"],row_d["type"],row_d["inr"],row_d["aed"],
                f"=D{r_num}+E{r_num}*{SET}!$B$4",
                row_d["cat"],row_d["freq"],row_d["ends"],0,row_d["notes"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r_num, j, value=v)
            cell.fill = fill("D6E4F0" if j==6 else bg); cell.border = border()
            if j==6:   cell.font=xf(); cell.number_format=FMT_INR; cell.alignment=ra()
            elif j in (4,5): cell.font=inp(); cell.number_format=FMT_INR; cell.alignment=ra()
            elif j==1: cell.font=bf(bold=True); cell.alignment=la()
            else:      cell.font=bf(sz=9); cell.alignment=ca() if j>1 else la()

    mc_total_r = 4 + len(mc_rows) + 1
    merge_sc(ws, mc_total_r, 1, 3, "TOTAL MONTHLY OUTFLOW", fnt=hf(sz=10), bg="F39C12", aln=ca())
    for col, fmt in [(4,FMT_INR),(5,FMT_AED),(6,FMT_INR)]:
        col_l = get_column_letter(col)
        cell = ws.cell(mc_total_r, col, value=f"=SUM({col_l}4:{col_l}{mc_total_r-1})")
        cell.font = hf(sz=10); cell.fill = fill("F39C12")
        cell.number_format = fmt; cell.alignment = ra(); cell.border = border()

    for col, w in [("A",30),("B",12),("C",12),("D",14),("E",12),("F",16),
                   ("G",12),("H",10),("I",12),("J",10),("K",35)]:
        cw(ws, col, w)

    # ── SHEET 7: Dashboard ────────────────────────────────────────
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1ABC9C"
    ASSETS_S = "'💰 Assets'"

    banner(ws, 1, f"📊  FINANCIAL DASHBOARD — {name}" + (f" & {spouse}" if spouse else ""), "1A2B4A", 6, sz=13)

    kpis = [
        ("💰 Total Assets",    f"={ASSETS_S}!G{gt_assets}", FMT_INR, "27AE60"),
        ("📋 Liabilities",     f"={LIAB}!F{liab_total_r}", FMT_INR, "E74C3C"),
        ("📈 Net Worth",       f"={ASSETS_S}!G{gt_assets}-{LIAB}!F{liab_total_r}", FMT_INR, "2D5986"),
        ("💸 Monthly Outflow", f"={MC}!F{mc_total_r}", FMT_INR, "F39C12"),
        ("🛡 Life Cover",      f"=IFERROR(SUMIF({POL}!G:G,\"Y\",{POL}!C:C),0)", FMT_INR, "8E44AD"),
        ("📅 As of",           "=TODAY()", "DD-MMM-YYYY", "1A2B4A"),
    ]
    for i, (label, formula, fmt, color) in enumerate(kpis):
        col = i + 1
        merge_sc(ws, 3, col, col, label, fnt=hf(sz=9), bg=color, aln=ca())
        cell = ws.cell(4, col, value=formula)
        cell.font = Font(name="Arial", size=13, bold=True, color=color)
        cell.fill = fill("FFFFFF"); cell.border = border()
        cell.number_format = fmt; cell.alignment = ca()
        rh(ws, 3, 26); rh(ws, 4, 40)
        ws.column_dimensions[get_column_letter(col)].width = 24

    # Asset breakdown
    section(ws, 6, "📊  ASSET BREAKDOWN", "1A2B4A", 5)
    hdr(ws, 7, ["Asset Class","Invested (Equiv)","Current Value (Equiv)","% of Total",""],start=1)
    for i, (label, _) in enumerate(asset_cats):
        r_num = 8 + i
        bg = "F2F2F2" if i % 2 == 1 else "FFFFFF"
        asset_row = data_start + i
        ws.cell(r_num,1).value = label
        ws.cell(r_num,1).font = bf(bold=True); ws.cell(r_num,1).fill = fill(bg)
        ws.cell(r_num,1).border = border(); ws.cell(r_num,1).alignment = la()
        for col, src_col, fmt in [(2,4,FMT_INR),(3,7,FMT_INR),(4,8,FMT_PCT)]:
            col_l = get_column_letter(src_col)
            cell = ws.cell(r_num, col, value=f"={ASSETS_S}!{col_l}{asset_row}")
            cell.font = xf(); cell.fill = fill(bg); cell.border = border()
            cell.number_format = fmt; cell.alignment = ra()

    # Liability breakdown
    liab_dash_r = 8 + len(asset_cats) + 2
    section(ws, liab_dash_r, "📋  LIABILITY BREAKDOWN", "E74C3C", 5)
    hdr(ws, liab_dash_r+1, ["Liability","Outstanding","EMI","Rate %",""], start=1)
    for i, liab in enumerate(liabilities):
        r_num = liab_dash_r + 2 + i
        bg = "F2F2F2" if i % 2 == 1 else "FFFFFF"
        liab_src = 4 + i
        ws.cell(r_num,1).value = f"={LIAB}!A{liab_src}"
        ws.cell(r_num,1).font = xf(); ws.cell(r_num,1).fill = fill(bg)
        ws.cell(r_num,1).border = border(); ws.cell(r_num,1).alignment = la()
        for col, liab_col, fmt in [(2,4,FMT_INR),(3,8,FMT_INR),(4,7,FMT_PCT)]:
            col_l = get_column_letter(liab_col)
            cell = ws.cell(r_num, col, value=f"={LIAB}!{col_l}{liab_src}")
            cell.font = xf(); cell.fill = fill(bg); cell.border = border()
            cell.number_format = fmt; cell.alignment = ra()

    # ── SHEET 8: Goal Tracker ─────────────────────────────────────
    ws = wb.create_sheet("🎯 Goal Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "F39C12"
    ws.freeze_panes = "A9"
    banner(ws, 1, "🎯  GOAL TRACKER — Your Financial Goals", "1A2B4A", 11)

    sc(ws,3,1,"Inflation Rate:",fnt=bf(bold=True),aln=ra())
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=3)
    sc(ws,3,4,0.06,fmt=FMT_PCT,fnt=inp(),bg="FEF9E7",aln=ra())
    sc(ws,4,1,"Expected Return:",fnt=bf(bold=True),aln=ra())
    ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=3)
    sc(ws,4,4,0.12,fmt=FMT_PCT,fnt=inp(),bg="FEF9E7",aln=ra())
    rh(ws,3,20); rh(ws,4,20)

    hdr(ws, 8, ["Goal","Priority","Target Year","Years Left","Present Value",
                "Target (Infl-adj)","Current Corpus","Grows To","SIP Required","Progress %","Notes"])

    GOAL_EMOJIS = {"Emergency Fund":"🆘","Retirement":"👴","Education":"🎓",
                   "Marriage":"💍","Property":"🏠","Holiday":"✈️","Car":"🚗","Other":"🎯"}
    PRIORITY_MAP = {"Emergency Fund":"🔴 Critical","Retirement":"🟡 Important",
                    "Education":"🟡 Important","Marriage":"🟢 Plan",
                    "Property":"🟢 Plan","Holiday":"🟢 Plan","Car":"🟢 Plan","Other":"🟢 Plan"}

    for i, goal in enumerate(goals):
        r_num = 9 + i
        bg = "F2F2F2" if r_num % 2 == 1 else "FFFFFF"
        gtype = goal.get("type","Other")
        label = f"{GOAL_EMOJIS.get(gtype,'🎯')} {goal.get('name', gtype)}"
        priority = PRIORITY_MAP.get(gtype,"🟢 Plan")
        yr = int(goal.get("targetYear", 2035))
        pv = sf(goal.get("amount", 0))
        p_bg = "FADBD8" if "🔴" in priority else ("FEF9E7" if "🟡" in priority else "D5F5E3")

        vals = [
            label, priority, yr,
            f"=C{r_num}-YEAR(TODAY())",
            pv,
            f"=E{r_num}*(1+$D$3)^D{r_num}",
            sf(goal.get("corpus",0)),
            f"=IF(G{r_num}>0,G{r_num}*(1+$D$4)^D{r_num},0)",
            f"=IFERROR(IF(D{r_num}<=0,0,(F{r_num}-H{r_num})*($D$4/12)/((1+$D$4/12)^(D{r_num}*12)-1)),0)",
            f"=IFERROR(G{r_num}/F{r_num},0)",
            goal.get("notes",""),
        ]
        fmts = ["","","","",FMT_INR,FMT_INR,FMT_INR,FMT_INR,FMT_INR,FMT_PCT,""]
        bgs  = [bg,p_bg,bg,"D6E4F0",bg,"D6E4F0","D6E4F0","D6E4F0","FDEBD0",bg,bg]
        for j, (v, fmt, cell_bg) in enumerate(zip(vals, fmts, bgs), 1):
            cell = ws.cell(r_num, j, value=v)
            cell.fill = fill(cell_bg); cell.border = border(); cell.number_format = fmt or "General"
            if j == 1:    cell.font = bf(bold=True); cell.alignment = la()
            elif j in (4,6,8,9): cell.font = xf(); cell.alignment = ra()
            elif j in (3,5,7): cell.font = inp(); cell.alignment = ra()
            elif j == 9:  cell.font = bf(bold=True, color="E74C3C"); cell.alignment = ra()
            elif j == 10: cell.font = bf(); cell.alignment = ca()
            else:         cell.font = bf(sz=9, bold=(j==2)); cell.alignment = ca() if j==2 else la()
        rh(ws, r_num, 22)

    sip_r = 9 + len(goals) + 1
    merge_sc(ws, sip_r, 1, 8, "TOTAL MONTHLY SIP REQUIRED", fnt=hf(sz=11), bg="E74C3C", aln=ca())
    cell = ws.cell(sip_r, 9, value=f"=SUM(I9:I{sip_r-1})")
    cell.font = hf(sz=12); cell.fill = fill("E74C3C")
    cell.number_format = FMT_INR; cell.alignment = ra(); cell.border = border()

    for col, w in [("A",30),("B",14),("C",12),("D",10),("E",16),("F",18),
                   ("G",16),("H",16),("I",18),("J",10),("K",40)]:
        cw(ws, col, w)

    # ── SHEET 9: Financial Health ─────────────────────────────────
    ws = wb.create_sheet("🩺 Financial Health")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "E74C3C"
    banner(ws, 1, "🩺  FINANCIAL HEALTH SCORECARD  (Max Score = 24)", "1A2B4A", 7)
    merge_sc(ws, 2, 1, 7, "Auto-calculated. Score: 3=Good  2=OK  1=Needs Work",
             fnt=bf(sz=9, color="555555"), aln=la())
    hdr(ws, 5, ["#","Metric","Your Value","Benchmark","Score","Verdict"])

    metrics = [
        ("1","EMI-to-Income Ratio",
         f"=IFERROR({LIAB}!F{liab_total_r}/{SET}!B7,0)", FMT_PCT,
         "< 35% Good | 35-50% OK | > 50% Bad",
         f"=IFERROR(IF({LIAB}!F{liab_total_r}/{SET}!B7<0.35,3,IF({LIAB}!F{liab_total_r}/{SET}!B7<0.5,2,1)),1)",
         f"=IFERROR(IF({LIAB}!F{liab_total_r}/{SET}!B7<0.35,\"✅ Good\",IF({LIAB}!F{liab_total_r}/{SET}!B7<0.5,\"⚠️ OK\",\"🔴 Bad\")),\"🔴 Bad\")"),
        ("2","Savings Ratio",
         f"=IFERROR(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F)/{SET}!B7,0)", FMT_PCT,
         "> 20% Good | 10-20% OK | < 10% Low",
         f"=IFERROR(IF(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F)/{SET}!B7>0.2,3,IF(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F)/{SET}!B7>0.1,2,1)),1)",
         f"=IFERROR(IF(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F)/{SET}!B7>0.2,\"✅ Good\",IF(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F)/{SET}!B7>0.1,\"⚠️ OK\",\"🔴 Low\")),\"🔴 Low\")"),
        ("3","Emergency Fund",
         f"=IFERROR(SUMIF({INV}!B:B,\"Savings\",{INV}!H:H),0)", FMT_INR,
         "> 3× EMI Good | 1-3× Building | < 1× Critical",
         f"=IFERROR(IF(SUMIF({INV}!B:B,\"Savings\",{INV}!H:H)>{LIAB}!F{liab_total_r}*3,3,IF(SUMIF({INV}!B:B,\"Savings\",{INV}!H:H)>100000,2,1)),1)",
         f"=IFERROR(IF(SUMIF({INV}!B:B,\"Savings\",{INV}!H:H)>{LIAB}!F{liab_total_r}*3,\"✅ Good\",IF(SUMIF({INV}!B:B,\"Savings\",{INV}!H:H)>100000,\"⚠️ Building\",\"🔴 Critical\")),\"🔴 Critical\")"),
        ("4","Term Insurance",
         f"=IFERROR(SUMIF({POL}!G:G,\"Y\",{POL}!C:C),0)", FMT_INR,
         "> 10× annual income",
         f"=IFERROR(IF(SUMIF({POL}!G:G,\"Y\",{POL}!C:C)>{SET}!B7*120,3,IF(SUMIF({POL}!G:G,\"Y\",{POL}!C:C)>{SET}!B7*60,2,1)),1)",
         f"=IFERROR(IF(SUMIF({POL}!G:G,\"Y\",{POL}!C:C)>{SET}!B7*120,\"✅ Adequate\",IF(SUMIF({POL}!G:G,\"Y\",{POL}!C:C)>{SET}!B7*60,\"⚠️ Partial\",\"🔴 Insufficient\")),\"🔴 Insufficient\")"),
        ("5","Equity Exposure",
         f"=IFERROR((SUMIF({INV}!B:B,\"Mutual Fund\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks India\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks UAE\",{INV}!H:H))/MAX(SUMIF({INV}!I:I,\"Active\",{INV}!H:H),1),0)", FMT_PCT,
         "> 30% Good | 15-30% OK | < 15% Low",
         f"=IFERROR(IF((SUMIF({INV}!B:B,\"Mutual Fund\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks India\",{INV}!H:H))/MAX(SUMIF({INV}!I:I,\"Active\",{INV}!H:H),1)>0.3,3,IF((SUMIF({INV}!B:B,\"Mutual Fund\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks India\",{INV}!H:H))/MAX(SUMIF({INV}!I:I,\"Active\",{INV}!H:H),1)>0.15,2,1)),1)",
         f"=IFERROR(IF((SUMIF({INV}!B:B,\"Mutual Fund\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks India\",{INV}!H:H))/MAX(SUMIF({INV}!I:I,\"Active\",{INV}!H:H),1)>0.3,\"✅ Good\",IF((SUMIF({INV}!B:B,\"Mutual Fund\",{INV}!H:H)+SUMIF({INV}!B:B,\"Stocks India\",{INV}!H:H))/MAX(SUMIF({INV}!I:I,\"Active\",{INV}!H:H),1)>0.15,\"⚠️ Low\",\"🔴 Very Low\")),\"🔴 Very Low\")"),
        ("6","Debt-to-Asset",
         f"=IFERROR({LIAB}!F{liab_total_r}/'💰 Assets'!G{gt_assets},0)", "0.00x",
         "< 0.5× Good | 0.5-1× Watch | > 1× Danger",
         f"=IFERROR(IF({LIAB}!F{liab_total_r}/'💰 Assets'!G{gt_assets}<0.5,3,IF({LIAB}!F{liab_total_r}/'💰 Assets'!G{gt_assets}<1,2,1)),1)",
         f"=IFERROR(IF({LIAB}!F{liab_total_r}/'💰 Assets'!G{gt_assets}<0.5,\"✅ Healthy\",IF({LIAB}!F{liab_total_r}/'💰 Assets'!G{gt_assets}<1,\"⚠️ Watch\",\"🔴 Danger\")),\"🔴 Danger\")"),
        ("7","Goal Progress",
         f"=IFERROR(AVERAGE('🎯 Goal Tracker'!J9:J{8+len(goals)}),0)", FMT_PCT,
         "> 50% Good | 20-50% Early | < 20% Starting",
         f"=IFERROR(IF(AVERAGE('🎯 Goal Tracker'!J9:J{8+len(goals)})>0.5,3,IF(AVERAGE('🎯 Goal Tracker'!J9:J{8+len(goals)})>0.2,2,1)),1)",
         f"=IFERROR(IF(AVERAGE('🎯 Goal Tracker'!J9:J{8+len(goals)})>0.5,\"✅ Good\",IF(AVERAGE('🎯 Goal Tracker'!J9:J{8+len(goals)})>0.2,\"⚠️ Early\",\"🔴 Starting\")),\"🔴 Starting\")"),
        ("8","Net Worth",
         f"=IFERROR('💰 Assets'!G{gt_assets}-{LIAB}!F{liab_total_r},0)", FMT_INR,
         "> 0 = Positive | < 0 = Liabilities exceed assets",
         f"=IFERROR(IF('💰 Assets'!G{gt_assets}-{LIAB}!F{liab_total_r}>0,3,2),1)",
         f"=IFERROR(IF('💰 Assets'!G{gt_assets}-{LIAB}!F{liab_total_r}>0,\"✅ Positive\",\"⚠️ Negative\"),\"⚠️ Negative\")"),
    ]

    for i, (num, metric, val_f, val_fmt, bench, score_f, verdict_f) in enumerate(metrics):
        r_num = 6 + i
        bg = "F2F2F2" if i % 2 == 1 else "FFFFFF"
        for j, (v, cell_bg, fnt_, aln_) in enumerate([
            (num, bg, bf(bold=True), ca()),
            (metric, bg, bf(bold=True), la()),
            (val_f, "D6E4F0", xf(), ra()),
            (bench, bg, bf(sz=8, color="555555"), la()),
            (score_f, "1A2B4A", bf(bold=True, color="FFFFFF"), ca()),
            (verdict_f, bg, bf(bold=True), ca()),
        ], 1):
            cell = ws.cell(r_num, j, value=v)
            cell.font = fnt_; cell.fill = fill(cell_bg)
            cell.border = border(); cell.alignment = aln_
            if j == 3: cell.number_format = val_fmt
        rh(ws, r_num, 22)

    ov_r = 6 + len(metrics)
    merge_sc(ws, ov_r, 1, 4, "OVERALL FINANCIAL HEALTH  (out of 24)", fnt=hf(sz=11), bg="4A235A", aln=ca())
    score_sum = "+".join([f"E{6+i}" for i in range(len(metrics))])
    ws.cell(ov_r,5).value = f"=IFERROR({score_sum},0)"
    ws.cell(ov_r,5).font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws.cell(ov_r,5).fill = fill("4A235A"); ws.cell(ov_r,5).alignment = ca(); ws.cell(ov_r,5).border = border()
    ws.cell(ov_r,6).value = f"=IF(E{ov_r}>=18,\"✅ Financially Healthy\",IF(E{ov_r}>=12,\"⚠️ Needs Work\",\"🔴 Take Action Now\"))"
    ws.cell(ov_r,6).font = hf(sz=11); ws.cell(ov_r,6).fill = fill("4A235A")
    ws.cell(ov_r,6).alignment = ca(); ws.cell(ov_r,6).border = border()
    rh(ws, ov_r, 28)
    for col, w in [("A",5),("B",26),("C",18),("D",35),("E",12),("F",22)]: cw(ws,col,w)

    # ── SHEET 10: Salary Allocation ───────────────────────────────
    ws = wb.create_sheet("💡 Salary Allocation")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "8E44AD"
    banner(ws, 1, "💡  IDEAL SALARY ALLOCATION PLANNER", "1A2B4A", 7)
    for r_num, label, val, fmt in [
        (4, "Primary Income:", f"={SET}!B5" if currency=="INR" else f"={SET}!B5*{SET}!B4", FMT_INR),
        (5, "Spouse Income:",  f"={SET}!B6" if currency=="INR" else f"={SET}!B6*{SET}!B4", FMT_INR),
        (6, "Combined:",       f"={SET}!B7", FMT_INR),
    ]:
        sc(ws, r_num, 1, label, fnt=bf(bold=True), aln=ra())
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=2)
        sc(ws, r_num, 3, val, fmt=fmt, fnt=xf(), bg="D6E4F0", aln=ra())
        rh(ws, r_num, 20)

    section(ws, 8, "ALLOCATION — Ideal % vs Actual", "2D5986", 7)
    hdr(ws, 9, ["#","Category","What's Included","Ideal %","Ideal ₹/month","Actual ₹/month","Gap"])
    alloc_cats = [
        ("1","Fixed EMIs & Rent","All loan EMIs + housing",0.35,
         f"=IFERROR(SUMIF({MC}!G:G,\"Liability\",{MC}!F:F),0)"),
        ("2","Living Expenses","Food, transport, utilities, kids",0.15,0),
        ("3","Investments","SIPs, PPF, NPS, gold",0.20,
         f"=IFERROR(SUMIF({MC}!G:G,\"Investment\",{MC}!F:F),0)"),
        ("4","Insurance Premiums","All policies",0.05,
         f"=IFERROR(SUMIF({MC}!G:G,\"Insurance\",{MC}!F:F),0)"),
        ("5","Emergency Fund","Liquid MF — till 6× EMI built",0.10,0),
        ("6","Charity / Family","Transfers, donations",0.05,
         f"=IFERROR(SUMIF({MC}!G:G,\"Charity\",{MC}!F:F),0)"),
        ("7","Buffer / Savings","Unspent → liquid MF",0.10,0),
    ]
    for i, (num, cat, inc, pct, actual) in enumerate(alloc_cats):
        r_num = 10 + i
        bg = "F2F2F2" if i % 2 == 1 else "FFFFFF"
        vals = [num, cat, inc, pct, f"={SET}!B7*D{r_num}", actual, f"=E{r_num}-F{r_num}"]
        fmts = ["","","",FMT_PCT,FMT_INR,FMT_INR,FMT_INR]
        bgs = [bg,bg,bg,"FEF9E7","D6E4F0","D6E4F0" if actual else "FEF9E7",bg]
        for j, (v, fmt, cell_bg) in enumerate(zip(vals,fmts,bgs),1):
            cell = ws.cell(r_num,j,value=v)
            cell.fill = fill(cell_bg); cell.border = border(); cell.number_format = fmt or "General"
            if j==4:   cell.font=inp(); cell.alignment=ca()
            elif j==5: cell.font=xf(); cell.alignment=ra()
            elif j==6: cell.font=(xf() if actual else inp()); cell.alignment=ra()
            elif j==7: cell.font=bf(); cell.alignment=ra()
            elif j==1: cell.font=bf(bold=True); cell.alignment=ca()
            elif j==2: cell.font=bf(bold=True); cell.alignment=la()
            else:      cell.font=bf(sz=9); cell.alignment=la()
        rh(ws, r_num, 22)

    sal_total = 10+len(alloc_cats)
    merge_sc(ws, sal_total, 1, 3, "TOTAL", fnt=hf(sz=10), bg="1A2B4A", aln=ca())
    for col, fmt in [(4,FMT_PCT),(5,FMT_INR),(6,FMT_INR),(7,FMT_INR)]:
        col_l = get_column_letter(col)
        cell = ws.cell(sal_total, col, value=f"=SUM({col_l}10:{col_l}{sal_total-1})")
        cell.font=hf(sz=10); cell.fill=fill("1A2B4A")
        cell.number_format=fmt; cell.alignment=ra(); cell.border=border()
    for col, w in [("A",5),("B",22),("C",35),("D",10),("E",18),("F",18),("G",16)]: cw(ws,col,w)

    # ── SHEET 11: Living Expenses ─────────────────────────────────
    ws = wb.create_sheet("💸 Living Expenses")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "F39C12"
    ws.freeze_panes = "B5"
    banner(ws, 1, "💸  MONTHLY LIVING EXPENSES LOG", "1A2B4A", 11)
    merge_sc(ws,2,1,11,"Enter spending each month. One row per month.",
             fnt=bf(sz=9,color="555555"),aln=la())
    exp_hdrs = ["Month","Household","Daily Living","Medical","Children",
                "Travel","Eating Out","Shopping","Education","Misc","TOTAL"]
    hdr(ws, 4, exp_hdrs); rh(ws,4,28)
    cur = datetime.now().replace(day=1)
    for i in range(24):
        month = cur.replace(year=cur.year+((cur.month+i-1)//12), month=((cur.month+i-1)%12)+1)
        r_num = 5+i; bg = "F2F2F2" if i%2==1 else "FFFFFF"
        sc(ws,r_num,1,month.strftime("%b-%y"),fnt=bf(bold=True),bg=bg,aln=ca())
        for j in range(2,11):
            cell=ws.cell(r_num,j,value=None)
            cell.font=inp(); cell.fill=fill("FEF9E7")
            cell.border=border(); cell.number_format=FMT_INR; cell.alignment=ra()
        cell=ws.cell(r_num,11,value=f"=SUM(B{r_num}:J{r_num})")
        cell.font=bf(bold=True); cell.fill=fill("D6E4F0")
        cell.border=border(); cell.number_format=FMT_INR; cell.alignment=ra()
        rh(ws,r_num,20)
    for col,w in [("A",10),("B",14),("C",14),("D",12),("E",12),("F",12),
                  ("G",12),("H",12),("I",12),("J",12),("K",14)]: cw(ws,col,w)

    # ── SHEET 12: Completed-Redeemed ─────────────────────────────
    ws = wb.create_sheet("✅ Completed-Redeemed")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "27AE60"
    banner(ws, 1, "✅  COMPLETED & REDEEMED INVESTMENTS", "1A2B4A", 9)
    hdr(ws, 3, ["Investment","By","Invested (INR)","Redemption Date",
                "Redemption Amount","Profit/Loss","CAGR %","XIRR %","Verdict"])
    for col,w in [("A",30),("B",10),("C",16),("D",14),("E",18),
                  ("F",16),("G",10),("H",10),("I",40)]: cw(ws,col,w)

    # ── SHEET 13: Investment Roadmap ─────────────────────────────
    ws = wb.create_sheet("📋 Investment Roadmap")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "1ABC9C"
    banner(ws, 1, f"📋  INVESTMENT ROADMAP — {name}" + (f" & {spouse}" if spouse else ""), "1A2B4A", 7, sz=13)
    merge_sc(ws,2,1,7,
        f"{country} | Risk: {risk} | Income: {income:,.0f} {currency} + Spouse: {spouse_inc:,.0f} {currency}",
        fnt=bf(sz=9,bold=True,color="E74C3C"),aln=la())

    section(ws, 4, "🚨  IMMEDIATE ACTIONS — Do These First", "E74C3C", 7)
    hdr(ws, 5, ["#","Action","Why","Amount","By When","Done?"])

    immediate = []
    has_term = any(p.get("type","")=="Term Insurance" and p.get("status","")=="Active" for p in policies)
    if not has_term:
        immediate.append(("1","Buy Term Insurance",
            f"No term cover found. Cover needed: ~{income*120:,.0f} (10× annual income)",
            f"~{income*0.005:,.0f}/month","This week","☐"))
    high_rate = max(liabilities, key=lambda x: sf(x.get("rate",0)), default=None) if liabilities else None
    if high_rate and sf(high_rate.get("rate",0)) > 12:
        immediate.append((str(len(immediate)+1),f"Clear {high_rate.get('name','')}",
            f"{high_rate.get('rate',0)}% interest — highest cost debt. Attack first.",
            f"₹{sf(high_rate.get('outstandingINR',0)):,.0f}","ASAP","☐"))
    immediate.append((str(len(immediate)+1),"Build Emergency Fund",
        f"Need 6× monthly expenses (~₹{income*6:,.0f}). Zero emergency fund = critical risk.",
        "₹10,000-25,000/month","12 months","☐"))

    for i, rd in enumerate(immediate):
        r_num = 6+i; bg = "FADBD8" if i%2==0 else "F5B7B1"
        for j,v in enumerate(rd,1):
            cell=ws.cell(r_num,j,value=v)
            cell.font=bf(sz=9,bold=(j==2)); cell.fill=fill(bg)
            cell.border=border(); cell.alignment=la()
        rh(ws,r_num,36)

    if liabilities:
        debt_r = 6+len(immediate)+2
        section(ws, debt_r, "🔵  DEBT PLAN — Pay in This Order", "2D5986", 7)
        hdr(ws, debt_r+1, ["Loan","Outstanding","Rate","EMI","Strategy","Close By"])
        sorted_liabs = sorted(liabilities, key=lambda x: sf(x.get("rate",0)), reverse=True)
        for i,liab in enumerate(sorted_liabs):
            r_num=debt_r+2+i; bg="D6EAF8" if i%2==0 else "EBF5FB"
            rate=sf(liab.get("rate",0))
            for j,v in enumerate([
                liab.get("name",""),f"₹{sf(liab.get('outstandingINR',0)):,.0f}",
                f"{rate}%",f"₹{sf(liab.get('emiINR',0)):,.0f}/mo",
                "Clear ASAP — high interest" if rate>15 else "Pay min only" if rate<5 else "Extra payments when possible",
                liab.get("closeBy",""),],1):
                cell=ws.cell(r_num,j,value=v)
                cell.font=bf(sz=9,bold=(j==1)); cell.fill=fill(bg)
                cell.border=border(); cell.alignment=la()
            rh(ws,r_num,36)

    inv_plan_r = (debt_r+2+len(liabilities)+2) if liabilities else (6+len(immediate)+2)
    section(ws, inv_plan_r, "🟢  INVESTMENT PLAN — Start in This Order", "27AE60", 7)
    hdr(ws, inv_plan_r+1, ["Priority","Investment","Instrument","Monthly","Goal","Timeline"])
    inv_plan = [
        ("🔴 #1","Emergency Fund","Liquid Mutual Fund",f"₹{min(income*0.1,25000):,.0f}","6× EMI","12 months"),
        ("🟡 #2","Term Insurance","HDFC Click2Protect / TATA AIA","Fixed premium","Life cover","Immediate"),
        ("🟡 #3","Health Insurance","Family floater ₹10L+","₹5,000-10,000/yr","Medical","Immediate"),
        ("🟢 #4","Index Fund SIP","Nifty 50 / Global ETF via IBKR",f"₹{min(income*0.1,10000):,.0f}","Wealth","After EF built"),
        ("🟢 #5","ELSS","Tax saving 80C","₹12,500","Tax + equity","When income allows"),
        ("🔵 #6","NPS","80CCD benefit","₹4,200+","Retirement","Ongoing"),
        ("🔵 #7","Children Goals","Index Fund / IBKR ETF","₹3,000-10,000","Education/Marriage","Start early"),
    ]
    for i,rd in enumerate(inv_plan):
        r_num=inv_plan_r+2+i; bg="D5F5E3" if i%2==0 else "E9F7EF"
        for j,v in enumerate(rd,1):
            cell=ws.cell(r_num,j,value=v)
            cell.font=bf(sz=9,bold=(j in (1,2))); cell.fill=fill(bg)
            cell.border=border(); cell.alignment=la()
        rh(ws,r_num,36)

    for col,w in [("A",10),("B",28),("C",28),("D",18),("E",24),("F",18),("G",10)]: cw(ws,col,w)

    # ── SHEET 14: Stock Watch ─────────────────────────────────────
    ws = wb.create_sheet("📉 Stock Watch")
    ws.sheet_view.showGridLines = False; ws.sheet_properties.tabColor = "1A2B4A"
    banner(ws, 1, "📉  STOCK WATCH LIST", "1A2B4A", 8)
    hdr(ws, 3, ["Stock/ETF","Exchange","Buy Price","Current Price","Qty",
                "Invested","Current Value","P&L","Return %","Notes"])
    for col,w in [("A",22),("B",10),("C",14),("D",14),("E",8),
                  ("F",14),("G",14),("H",14),("I",10),("J",30)]: cw(ws,col,w)

    # Desired sheet order
    desired = ["⚙ Settings","📈 Investments","📋 Liabilities","💰 Assets",
               "🛡 Policies","📅 Monthly Commitments","📊 Dashboard",
               "🎯 Goal Tracker","🩺 Financial Health","📋 Investment Roadmap",
               "💡 Salary Allocation","💸 Living Expenses","📉 Stock Watch","✅ Completed-Redeemed"]
    for i, sname in enumerate(desired):
        if sname in wb.sheetnames:
            idx = wb.sheetnames.index(sname)
            if idx != i:
                wb.move_sheet(sname, offset=-(idx-i))

    return wb


@app.route("/generate", methods=["POST"])
def generate():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
        wb = build_tracker(data)
        output = io.BytesIO()
        wb.save(output); output.seek(0)
        fname = f"Financial_Tracker_{data.get('name','User').replace(' ','_')}.xlsx"
        return send_file(output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=fname)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

if __name__ == "__main__":
    import os
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port, debug=False)
